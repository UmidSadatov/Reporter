from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pyexcel as p


def convert_to_xlsx(filename: str):
    if filename.split('.')[-1] == 'xls':
        p.save_book_as(file_name=filename,
                       dest_file_name=filename + 'x')
        return filename + 'x'


class ExcelBook:
    def __init__(self, filename: str):
        extension = filename.split('.')[-1].lower()

        if extension not in ['xlsx', 'xls']:
            raise AttributeError('File extension must be .xlsx or .xls')
        elif extension == 'xls':
            self.filename = convert_to_xlsx(filename)
        else:
            self.filename = filename
        self.book = load_workbook(self.filename, data_only=True)
        self.number_of_sheets = len(self.book.worksheets)

    def get_cell_value(self, cell: str, sheet_index=0, round_value=True):

        try:
            sheet = self.book.worksheets[sheet_index]
        except IndexError:
            return None

        # value = sheet[cell].value if sheet[cell].value is not None else 0
        #
        # if type(value) is float and round_value:
        #     value = round(value, 2)

        return sheet[cell].value

    def set_cell_value(self, cell, value, sheet_index=0):
        sheet = self.book.worksheets[sheet_index]
        sheet[cell].value = value

    def get_max_row(self, sheet_index=0):
        sheet = self.book.worksheets[sheet_index]
        return sheet.max_row


print(ExcelBook("files/Интеллия.xlsx").get_cell_value('K2'))